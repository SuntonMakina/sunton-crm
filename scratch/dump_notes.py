import openpyxl

excel_path = '/Users/berkhan/Downloads/2026 - Mayıs Haziran Verileri.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Lead Takip']

headers = [str(cell.value).strip() if cell.value is not None else f"Col{i+1}" for i, cell in enumerate(sheet[2])]

rows = []
for r in range(3, sheet.max_row + 1):
    lead_id = sheet.cell(row=r, column=1).value
    if not lead_id or str(lead_id).strip() == "":
        continue
    row_data = {h: sheet.cell(row=r, column=i+1).value for i, h in enumerate(headers)}
    rows.append((r, lead_id, row_data))

with open('/Users/berkhan/Documents/Websiteler/Call Center CRM/scratch/notes_analysis.txt', 'w', encoding='utf-8') as f:
    for r, lid, rdata in rows:
        note = rdata.get('İlk Mesaj / Arama Notu')
        summary = rdata.get('Görüşme Özeti / Sonuç')
        status = rdata.get('Lead Durumu')
        rep = rdata.get('Satış Uzmanı')
        konusma = rdata.get('Konuşma Yapıldı mı?')
        olumlu = rdata.get('Dönüş Olumlu mu?')
        f.write(f"Row {r} | ID {lid} | Status: {status} | Rep: {rep} | Call: {konusma} | Olumlu: {olumlu}\n")
        f.write(f"  Note: {note}\n")
        f.write(f"  Summary: {summary}\n")
        f.write("-" * 80 + "\n")
print("Written notes_analysis.txt")
