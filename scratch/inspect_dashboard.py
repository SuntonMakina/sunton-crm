import openpyxl

excel_path = '/Users/berkhan/Downloads/2026 - Mayıs Haziran Verileri.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=False) # Load with formulas
sheet = wb['Dashboard']

for r in range(1, 100):
    row_vals = []
    for c in range(1, 30):
        val = sheet.cell(row=r, column=c).value
        if val is not None:
            row_vals.append(f"Col{c}({openpyxl.utils.get_column_letter(c)}{r}): {val}")
    if row_vals:
        print(f"Row {r}: " + " | ".join(row_vals))
