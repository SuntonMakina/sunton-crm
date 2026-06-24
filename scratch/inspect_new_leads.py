import openpyxl

excel_path = '/Users/berkhan/Downloads/2026 - Mayıs Haziran Verileri.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Lead Takip']

headers = [str(cell.value).strip() if cell.value is not None else f"Col{i+1}" for i, cell in enumerate(sheet[2])]

# Let's inspect a few rows that have status resolved to "Yeni Lead" in previous logic
# and see if there are other columns that have content.
rows_with_data = []
for r in range(3, sheet.max_row + 1):
    lead_id = sheet.cell(row=r, column=1).value
    if not lead_id or str(lead_id).strip() == "":
        continue
        
    row_data = {h: sheet.cell(row=r, column=i+1).value for i, h in enumerate(headers)}
    
    # Check what our previous logic would resolve:
    status_raw = row_data.get('Lead Durumu')
    olumlu_raw = row_data.get('Dönüş Olumlu mu?')
    konusma_raw = row_data.get('Konuşma Yapıldı mı?')
    rep_raw = row_data.get('Satış Uzmanı')
    rep = str(rep_raw).strip() if rep_raw else "-"
    follow_up_raw = row_data.get('Sonraki Takip Tarihi')
    
    # Previous logic status determination:
    status_text = 'Yeni Lead'
    if status_raw == 'Olumsuz' or olumlu_raw == 'Hayır':
        status_text = 'İlgileniyor'
    elif konusma_raw == 'Hayır':
        status_text = 'Ulaşılamadı'
    elif rep and rep != "-":
        status_text = 'Satış Uzmanına İletildi'
    elif konusma_raw == 'Evet':
        status_text = 'Görüşme Yapıldı'
    elif follow_up_raw:
        status_text = 'Geri Aranacak'
        
    if status_text == 'Yeni Lead':
        # Let's check what fields are not empty in this row
        non_empty = {k: v for k, v in row_data.items() if v is not None and str(v).strip() not in ('', '-')}
        rows_with_data.append((r, lead_id, non_empty))

print(f"Total leads resolving to 'Yeni Lead': {len(rows_with_data)}")
print("\n--- Example 10 leads with non-empty details ---")
for r, lid, details in rows_with_data[:10]:
    print(f"\nRow {r} (ID: {lid}):")
    for k, v in details.items():
        if k not in ('Lead ID', 'İlk Temas Tarihi', 'İlk Temas Saati', 'Ad Soyad / Firma', 'Telefon Numarası', 'Şehir', 'İletişim Kanalı', 'Reklam Kaynağı', 'İstenen Makine / Ürün', 'Öncelik'):
            print(f"  {k}: {v}")
