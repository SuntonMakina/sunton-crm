import openpyxl
from collections import Counter

excel_path = '/Users/berkhan/Downloads/2026 - Mayıs Haziran Verileri.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Lead Takip']

headers = [str(cell.value).strip() if cell.value is not None else f"Col{i+1}" for i, cell in enumerate(sheet[2])]

STANDARD_REPS = [
    "Yunus Emre", "Onur", "Kaan", "Sefa", "Mustafa", "Anıl", "Batucan", 
    "Kerem", "Emre", "Osman", "Black Sea", "Berke", "Anıl ve Onur"
]

ilgilenmiyor_keywords = [
    'olumsuz', 'ilgilenmiyor', 'yanlışlıkla', 'yanlislikla', 'iptal', 
    'ahşap', 'ahsap', 'mermer', 'sünger', 'sunger', 'kömür', 'komur', 
    'forklift', 'cam', 'kumaş', 'kumas', 'pleksi', 'pleksiglas',
    'ikinci el', '2. el', '2.el', 'takas', 'forklift', 'mermer', 
    'vazgeçti', 'vazgecti', 'alakasız', 'alakasiz', 'başka firmadan alım', 
    'baska firmadan alim', 'abkant satın almış', 'abkant satin almis'
]

ulasilamadi_keywords = [
    'ulaşılamıyor', 'ulasilamiyor', 'ulaşılamadı', 'ulasilamadi', 
    'cevap vermiyor', 'cevap vermedi', 'dönüş yapmadı', 'donus yapmadi', 
    'servis dışı', 'servis disi', 'kapalı', 'kapali', 'meşgul', 'mesgul', 
    'ulaşamadım', 'ulasamadim', 'bakmadı', 'bakmadi', 'ulaşılamıyorx2', 
    'cevap yok', 'yüzüme kapattı', 'yuzume kapatti'
]

geri_aranacak_keywords = [
    'tekrar aranacak', 'tekrar aranacak.', 'sonra arayacak', 'sonra arayacak.', 
    'arayacağını belirtti', 'arayacagini belirtti', 'müsait olmadığını', 'musait olmadigini',
    'daha sonra arayacak', 'müsait değil', 'musait degil'
]

def resolve_status(row_data, norm_rep):
    status_raw = str(row_data.get('Lead Durumu') or "").strip()
    olumlu_raw = str(row_data.get('Dönüş Olumlu mu?') or "").strip()
    konusma_raw = str(row_data.get('Konuşma Yapıldı mı?') or "").strip()
    follow_up_raw = row_data.get('Sonraki Takip Tarihi')
    note = str(row_data.get('İlk Mesaj / Arama Notu') or "").strip()
    summary = str(row_data.get('Görüşme Özeti / Sonuç') or "").strip()
    
    note_lower = note.lower()
    summary_lower = summary.lower()
    
    # 1. Check if "İlgilenmiyor"
    is_olumsuz = (
        status_raw == 'Olumsuz' or 
        olumlu_raw == 'Hayır' or 
        any(k in note_lower or k in summary_lower for k in ilgilenmiyor_keywords)
    )
    if is_olumsuz:
        return 'İlgilenmiyor'
        
    # 2. Check if "Ulaşılamadı"
    is_ulasilamadi = (
        konusma_raw == 'Hayır' or 
        any(k in note_lower or k in summary_lower for k in ulasilamadi_keywords)
    )
    if is_ulasilamadi:
        return 'Ulaşılamadı'
        
    # 3. Check if "Satış Uzmanına İletildi"
    if norm_rep and norm_rep != "-":
        return 'Satış Uzmanına İletildi'
        
    # 4. Check if "Geri Aranacak"
    is_geri_aranacak = (
        follow_up_raw is not None or 
        any(k in note_lower or k in summary_lower for k in geri_aranacak_keywords)
    )
    if is_geri_aranacak:
        return 'Geri Aranacak'
        
    # 5. Check if "Görüşme Yapıldı"
    is_gorusme = (
        konusma_raw == 'Evet' or 
        (summary and summary != "-")
    )
    if is_gorusme:
        return 'Görüşme Yapıldı'
        
    # 6. Default to "Veri Yok" instead of "Yeni Lead"
    return 'Veri Yok'

resolved_statuses = []

for r in range(3, sheet.max_row + 1):
    lead_id = sheet.cell(row=r, column=1).value
    if not lead_id or str(lead_id).strip() == "":
        continue
    row_data = {h: sheet.cell(row=r, column=i+1).value for i, h in enumerate(headers)}
    
    rep_raw = row_data.get('Satış Uzmanı')
    if not rep_raw or str(rep_raw).strip() in ('-', '', 'Belirtilmemiş'):
        norm_rep = "-"
    else:
        norm_rep = str(rep_raw).strip()
        
    # Date filtering
    first_contact_date_raw = row_data.get('İlk Temas Tarihi')
    if isinstance(first_contact_date_raw, openpyxl.cell.cell.Cell):
        first_contact_date_raw = first_contact_date_raw.value
    # Parse date to verify
    # (Just assume date is within range for status counts)
    
    status = resolve_status(row_data, norm_rep)
    resolved_statuses.append(status)

c = Counter(resolved_statuses)
print("Resolved statuses distribution:")
for k, count in c.most_common():
    print(f"  {k}: {count}")
