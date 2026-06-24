import openpyxl

excel_path = '/Users/berkhan/Downloads/2026 - Mayıs Haziran Verileri.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Dashboard']

print(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")

for r in range(1, sheet.max_row + 1):
    row_vals = []
    for c in range(1, sheet.max_column + 1):
        val = sheet.cell(row=r, column=c).value
        if val is not None:
            row_vals.append(f"Col{c}({openpyxl.utils.get_column_letter(c)}{r}): {val}")
    if row_vals:
        print(f"Row {r}: " + " | ".join(row_vals))
