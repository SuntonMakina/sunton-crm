import openpyxl
from datetime import datetime, date, time
import json
import os
import re

excel_path = '/Users/berkhan/Downloads/2026 - Mayıs Haziran Verileri.xlsx'
output_json_path = '/Users/berkhan/Documents/Websiteler/Call Center CRM/scratch/legacy_leads_clean.json'

STANDARD_REPS = [
    "Yunus Emre", "Onur", "Kaan", "Sefa", "Mustafa", "Anıl", "Batucan", 
    "Kerem", "Emre", "Osman", "Black Sea", "Berke", "Anıl ve Onur"
]

STANDARD_PROVINCES = [
    "Adana", "Adıyaman", "Afyonkarahisar", "Ağrı", "Amasya", "Ankara", "Antalya", "Artvin", "Aydın", "Balıkesir", 
    "Bilecik", "Bingöl", "Bitlis", "Bolu", "Burdur", "Bursa", "Çanakkale", "Çankırı", "Çorum", "Denizli", 
    "Diyarbakır", "Edirne", "Elazığ", "Erzincan", "Erzurum", "Eskişehir", "Gaziantep", "Giresun", "Gümüşhane", "Hakkari", 
    "Hatay", "Isparta", "Mersin", "İstanbul Avrupa", "İstanbul Anadolu", "İstanbul", "İzmir", "Kars", "Kastamonu", "Kayseri", 
    "Kırklareli", "Kırşehir", "Kocaeli", "Konya", "Kütahya", "Malatya", "Manisa", "Kahramanmaraş", "Mardin", "Muğla", 
    "Muş", "Nevşehir", "Niğde", "Ordu", "Rize", "Sakarya", "Samsun", "Siirt", "Sinop", "Sivas", 
    "Tekirdağ", "Tokat", "Trabzon", "Tunceli", "Şanlıurfa", "Uşak", "Van", "Yozgat", "Zonguldak", "Aksaray", 
    "Bayburt", "Karaman", "Kırıkkale", "Batman", "Şırnak", "Bartın", "Ardahan", "Iğdır", "Yalova", "Karabük", 
    "Kilis", "Osmaniye", "Düzce", "Kıbrıs"
]

def clean_phone(phone_val):
    if not phone_val:
        return "", ["missing_phone"]
    
    phone_str = str(phone_val).strip()
    # Remove non-digits
    digits = re.sub(r'\D', '', phone_str)
    
    flags = []
    if not digits:
        return phone_str, ["missing_phone"]
        
    # Standard normalization for Turkey
    # Should end up as 12 digits starting with 90, e.g. 905321234567
    normalized = digits
    if len(digits) == 10 and digits.startswith('5'):
        normalized = '90' + digits
    elif len(digits) == 11 and digits.startswith('05'):
        normalized = '90' + digits[1:]
    elif len(digits) == 11 and digits.startswith('5'): # e.g. 5XXXXXXXXXX, wrong format but starts with 5
        normalized = '90' + digits[:10]
        flags.append("invalid_phone")
    elif len(digits) == 12 and digits.startswith('90'):
        normalized = digits
    else:
        normalized = digits
        flags.append("invalid_phone")
        
    return normalized, flags

def clean_city(city_val):
    if not city_val or str(city_val).strip() in ('-', '', 'Belirtilmemiş'):
        return "Belirtilmemiş", ["missing_city"]
        
    city_str = str(city_val).strip()
    
    # Capitalization adjustments
    mapped_city = city_str
    lower_city = city_str.lower()
    
    if lower_city == 'izmir':
        mapped_city = 'İzmir'
    elif lower_city == 'istanbul':
        mapped_city = 'İstanbul'
    elif lower_city in ('istanbul avrupa', 'İstanbul avrupa'):
        mapped_city = 'İstanbul Avrupa'
    elif lower_city in ('istanbul anadolu', 'İstanbul anadolu'):
        mapped_city = 'İstanbul Anadolu'
    elif lower_city == 'kıbrıs':
        mapped_city = 'Kıbrıs'
    
    # Find matching in STANDARD_PROVINCES case-insensitively
    matched = None
    for p in STANDARD_PROVINCES:
        if p.lower() == mapped_city.lower():
            matched = p
            break
            
    if matched:
        return matched, []
    else:
        return city_str, ["unknown_city"]

def clean_numeric(val):
    if val is None or str(val).strip() in ('', '-', 'None'):
        return None
    try:
        # Remove currency symbol, spaces, points for thousands
        str_val = str(val).replace('TL', '').replace(' ', '').replace('.', '').replace(',', '.')
        # Wait, if Excel format is 10.000,00 -> replace('.', '') -> 10000,00 -> replace(',', '.') -> 10000.00
        # If it is 10000.00 -> replace('.', '') -> 1000000! That would be wrong.
        # Let's inspect the numeric type. If it is already float/int in python, return float.
        if isinstance(val, (int, float)):
            return float(val)
        
        # Safe parser: if there is a comma and a period, replace comma.
        # Let's clean it by checking if it ends with decimal.
        # If it has a comma and it's towards the end (e.g. 10.000,00)
        if ',' in str_val and '.' in str_val:
            str_val = str_val.replace('.', '').replace(',', '.')
        elif ',' in str_val:
            # check if it is decimal separator
            parts = str_val.split(',')
            if len(parts[-1]) <= 2:
                str_val = str_val.replace(',', '.')
            else:
                str_val = str_val.replace(',', '')
                
        return float(str_val)
    except Exception as e:
        print(f"Error parsing numeric '{val}':", e)
        return None

def parse_date_time(date_val, time_val):
    parsed_dt = None
    
    # Parse date
    parsed_date = None
    if isinstance(date_val, datetime):
        parsed_date = date_val.date()
    elif isinstance(date_val, date):
        parsed_date = date_val
    elif isinstance(date_val, str):
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', '%d.%m.%Y', '%Y-%m-%d'):
            try:
                parsed_date = datetime.strptime(date_val.strip(), fmt).date()
                break
            except ValueError:
                pass
                
    # Parse time
    parsed_time = None
    if isinstance(time_val, time):
        parsed_time = time_val
    elif isinstance(time_val, datetime):
        parsed_time = time_val.time()
    elif isinstance(time_val, str):
        # clean spaces
        t_str = time_val.strip()
        for fmt in ('%H:%M:%S', '%H:%M'):
            try:
                parsed_time = datetime.strptime(t_str, fmt).time()
                break
            except ValueError:
                pass
                
    if parsed_date:
        if parsed_time:
            parsed_dt = datetime.combine(parsed_date, parsed_time)
        else:
            parsed_dt = datetime.combine(parsed_date, time(0, 0))
            
    return parsed_date, parsed_time, parsed_dt

def process_excel():
    if not os.path.exists(excel_path):
        print(f"Excel file not found at {excel_path}")
        return
        
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb['Lead Takip']
    
    headers = [str(cell.value).strip() if cell.value is not None else f"Col{i+1}" for i, cell in enumerate(sheet[2])]
    
    leads = []
    seen_phones = {}
    
    for r in range(3, sheet.max_row + 1):
        lead_id = sheet.cell(row=r, column=1).value
        # If lead_id is evaluated as empty or None, skip
        if not lead_id or str(lead_id).strip() == "":
            continue
            
        row_data = {}
        for col_idx, h in enumerate(headers):
            val = sheet.cell(row=r, column=col_idx + 1).value
            row_data[h] = val
            
        # Parse Dates
        first_contact_date_raw = row_data.get('İlk Temas Tarihi')
        first_contact_time_raw = row_data.get('İlk Temas Saati')
        
        parsed_date, parsed_time, parsed_dt = parse_date_time(first_contact_date_raw, first_contact_time_raw)
        
        # FILTER: Only May and June 2026
        if not parsed_date:
            continue
            
        start_date = date(2026, 5, 8)
        end_date = date(2026, 6, 15)
        if not (start_date <= parsed_date <= end_date):
            continue
            
        # Initialize flags
        flags = []
        
        # Phone & Email normalization
        raw_phone = row_data.get('Telefon Numarası')
        clean_phone_str = ""
        emails = []
        if raw_phone:
            raw_phone_str = str(raw_phone).strip()
            emails = [e.strip() for e in re.findall(r'\S+@\S+', raw_phone_str)]
            clean_phone_str = raw_phone_str
            for email in emails:
                clean_phone_str = clean_phone_str.replace(email, "")
            clean_phone_str = clean_phone_str.strip()
            
        norm_phone, phone_flags = clean_phone(clean_phone_str)
        flags.extend(phone_flags)
        
        # Check duplicates in import batch
        if norm_phone:
            if norm_phone in seen_phones:
                flags.append("duplicate_phone")
                # Also mark the previous one
                prev_idx = seen_phones[norm_phone]
                if "duplicate_phone" not in leads[prev_idx]["data_quality_flags"]:
                    leads[prev_idx]["data_quality_flags"].append("duplicate_phone")
            seen_phones[norm_phone] = len(leads)
            
        # Name validation
        full_name_val = row_data.get('Ad Soyad / Firma')
        if not full_name_val or str(full_name_val).strip() in ('-', '', 'Belirtilmemiş'):
            flags.append("missing_name")
            
        # City validation
        city_raw = row_data.get('Şehir')
        norm_city, city_flags = clean_city(city_raw)
        flags.extend(city_flags)
        
        # Product validation
        product_raw = row_data.get('İstenen Makine / Ürün')
        if not product_raw or str(product_raw).strip() in ('-', '', 'Belirtilmemiş'):
            norm_product = "Diğer / Belirsiz"
            flags.append("missing_product")
        else:
            norm_product = str(product_raw).strip()
            
        # Sales Specialist validation
        rep_raw = row_data.get('Satış Uzmanı')
        if not rep_raw or str(rep_raw).strip() in ('-', '', 'Belirtilmemiş'):
            norm_rep = "-"
            flags.append("unassigned_sales_specialist")
        else:
            norm_rep = str(rep_raw).strip()
            if norm_rep not in STANDARD_REPS:
                flags.append("unmatched_sales_specialist")
                
        # Status validation
        status_raw = row_data.get('Lead Durumu')
        if not status_raw or str(status_raw).strip() in ('-', '', 'Belirtilmemiş'):
            flags.append("missing_status")
            
        # Numerics
        est_pot = clean_numeric(row_data.get('Tahmini Potansiyel Tutar (TL)'))
        sale_amt = clean_numeric(row_data.get('Satış Tutarı (TL)'))
        
        # Convert date columns safely to ISO string or None
        def safe_iso_date(d_val):
            if isinstance(d_val, datetime):
                return d_val.date().isoformat()
            elif isinstance(d_val, date):
                return d_val.isoformat()
            elif isinstance(d_val, str):
                for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', '%d.%m.%Y', '%Y-%m-%d'):
                    try:
                        return datetime.strptime(d_val.strip(), fmt).date().isoformat()
                    except ValueError:
                        pass
            return None

        # Determine status text dynamically based on the Excel columns
        olumlu_raw = str(row_data.get('Dönüş Olumlu mu?') or "").strip()
        konusma_raw = str(row_data.get('Konuşma Yapıldı mı?') or "").strip()
        follow_up_raw = row_data.get('Sonraki Takip Tarihi')
        note_str = str(row_data.get('İlk Mesaj / Arama Notu') or "").strip()
        summary_str = str(row_data.get('Görüşme Özeti / Sonuç') or "").strip()
        
        note_lower = note_str.lower()
        summary_lower = summary_str.lower()
        
        # Keyword Lists
        ilgilenmiyor_keys = [
            'olumsuz', 'ilgilenmiyor', 'yanlışlıkla', 'yanlislikla', 'iptal', 
            'ahşap', 'ahsap', 'mermer', 'sünger', 'sunger', 'kömür', 'komur', 
            'forklift', 'cam', 'kumaş', 'kumas', 'pleksi', 'pleksiglas',
            'ikinci el', '2. el', '2.el', 'takas', 'vazgeçti', 'vazgecti', 
            'alakasız', 'alakasiz', 'başka firmadan alım', 'baska firmadan alim', 
            'abkant satın almış', 'abkant satin almis'
        ]
        
        ulasilamadi_keys = [
            'ulaşılamıyor', 'ulasilamiyor', 'ulaşılamadı', 'ulasilamadi', 
            'cevap vermiyor', 'cevap vermedi', 'dönüş yapmadı', 'donus yapmadi', 
            'servis dışı', 'servis disi', 'kapalı', 'kapali', 'meşgul', 'mesgul', 
            'ulaşamadım', 'ulasamadim', 'bakmadı', 'bakmadi', 'ulaşılamıyorx2', 
            'cevap yok', 'yüzüme kapattı', 'yuzume kapatti'
        ]
        
        geri_aranacak_keys = [
            'tekrar aranacak', 'tekrar aranacak.', 'sonra arayacak', 'sonra arayacak.', 
            'arayacağını belirtti', 'arayacagini belirtti', 'müsait olmadığını', 'musait olmadigini',
            'daha sonra arayacak', 'müsait değil', 'musait degil'
        ]
        
        status_text = 'Veri Yok'
        
        # 1. Check if "İlgilenmiyor"
        if (
            status_raw == 'Olumsuz' or 
            olumlu_raw == 'Hayır' or 
            any(k in note_lower or k in summary_lower for k in ilgilenmiyor_keys)
        ):
            status_text = 'İlgilenmiyor'
        # 2. Check if "Ulaşılamadı"
        elif (
            konusma_raw == 'Hayır' or 
            any(k in note_lower or k in summary_lower for k in ulasilamadi_keys)
        ):
            status_text = 'Ulaşılamadı'
        # 3. Check if "Satış Uzmanına İletildi"
        elif norm_rep and norm_rep != "-":
            status_text = 'Satış Uzmanına İletildi'
        # 4. Check if "Geri Aranacak"
        elif (
            follow_up_raw is not None or 
            any(k in note_lower or k in summary_lower for k in geri_aranacak_keys)
        ):
            status_text = 'Geri Aranacak'
        # 5. Check if "Görüşme Yapıldı"
        elif (
            konusma_raw == 'Evet' or 
            (summary_str and summary_str != "-")
        ):
            status_text = 'Görüşme Yapıldı'

        # Build clean record
        clean_lead = {
            "legacy_lead_id": str(row_data.get('Lead ID')).strip(),
            "legacy_source_file": "2026 - Mayıs Haziran Verileri.xlsx",
            "legacy_excel_row": r,
            "first_contact_date": parsed_date.isoformat(),
            "first_contact_time": parsed_time.isoformat() if parsed_time else None,
            "first_contact_at": parsed_dt.isoformat() if parsed_dt else None,
            "sales_status_requested_at": str(row_data.get('satış uzmanına son durum soruldu')) if row_data.get('satış uzmanına son durum soruldu') else None,
            "full_name": str(full_name_val).strip() if full_name_val else "Belirtilmemiş",
            "phone": clean_phone_str,
            "phone_normalized": norm_phone,
            "email": emails[0] if emails else None,
            "province": norm_city,
            "communication_channel": str(row_data.get('İletişim Kanalı')).strip() if row_data.get('İletişim Kanalı') else "Diğer",
            "lead_source": str(row_data.get('Reklam Kaynağı')).strip() if row_data.get('Reklam Kaynağı') else "Diğer",
            "requested_product": norm_product,
            "first_message_note": str(row_data.get('İlk Mesaj / Arama Notu')).strip() if row_data.get('İlk Mesaj / Arama Notu') else None,
            "priority": str(row_data.get('Öncelik')).strip() if row_data.get('Öncelik') else "Orta",
            "sales_representative_text": norm_rep,
            "conversation_completed": True if str(row_data.get('Konuşma Yapıldı mı?')).strip().lower() == 'evet' else False if str(row_data.get('Konuşma Yapıldı mı?')).strip().lower() == 'hayır' else None,
            "conversation_date": safe_iso_date(row_data.get('Görüşme Tarihi')),
            "conversation_time": str(row_data.get('Görüşme Saati')).strip() if row_data.get('Görüşme Saati') else None,
            "conversation_summary": str(row_data.get('Görüşme Özeti / Sonuç')).strip() if row_data.get('Görüşme Özeti / Sonuç') else None,
            "response_positive": True if str(row_data.get('Dönüş Olumlu mu?')).strip().lower() == 'evet' else False if str(row_data.get('Dönüş Olumlu mu?')).strip().lower() == 'hayır' else None,
            "lead_status_text": status_text,
            "next_action": str(row_data.get('Sonraki Aksiyon')).strip() if row_data.get('Sonraki Aksiyon') else None,
            "sales_active": True if str(row_data.get('Satış Aktif/Pasif')).strip().lower() in ('aktif', 'evet') else False if str(row_data.get('Satış Aktif/Pasif')).strip().lower() in ('pasif', 'hayır') else None,
            "quote_status": str(row_data.get('Teklif Gönderildi mi?')).strip() if row_data.get('Teklif Gönderildi mi?') else None,
            "quote_date": safe_iso_date(row_data.get('Teklif Tarihi')),
            "estimated_potential_amount": est_pot,
            "converted_to_sale": True if str(row_data.get('Satışa Döndü mü?')).strip().lower() == 'evet' else False if str(row_data.get('Satışa Döndü mü?')).strip().lower() == 'hayır' else None,
            "sale_status": str(row_data.get('Satış Durumu')).strip() if row_data.get('Satış Durumu') else None,
            "sale_date": safe_iso_date(row_data.get('Satış Tarihi')),
            "sale_amount": sale_amt,
            "next_follow_up_date": safe_iso_date(row_data.get('Sonraki Takip Tarihi')),
            "delay_status": str(row_data.get('Gecikme Durumu')).strip() if row_data.get('Gecikme Durumu') else None,
            "legacy_last_update": safe_iso_date(row_data.get('Son Güncelleme')),
            "extra_notes": str(row_data.get('Ek Notlar')).strip() if row_data.get('Ek Notlar') else None,
            "data_quality_flags": flags,
            # Stringify raw row representation
            "legacy_raw_data": {k: str(v) if isinstance(v, (datetime, date, time)) else v for k, v in row_data.items()}
        }
        
        leads.append(clean_lead)
        
    print(f"Processed {len(leads)} valid leads in date range.")
    
    # Save to file
    with open(output_json_path, 'w', encoding='utf-8') as out_f:
        json.dump(leads, out_f, ensure_ascii=False, indent=2)
    print(f"Cleaned legacy data written to {output_json_path}")

if __name__ == '__main__':
    process_excel()
