import openpyxl

excel_path = '/Users/berkhan/Downloads/2026 - Mayıs Haziran Verileri.xlsx'
wb = openpyxl.load_workbook(excel_path, read_only=True)
print("Sheet names:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\nSheet: {sheet_name}")
    # Print first 2 rows
    rows = list(sheet.iter_rows(max_row=2, values_only=True))
    if len(rows) > 0:
        print("Header row:", [r for r in rows[0] if r is not None][:15])
        if len(rows) > 1:
            print("First data row:", [r for r in rows[1] if r is not None][:15])
