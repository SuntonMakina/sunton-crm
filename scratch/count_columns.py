import openpyxl

excel_path = '/Users/berkhan/Downloads/2026 - Mayıs Haziran Verileri.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Lead Takip']

headers = [str(cell.value).strip() if cell.value is not None else f"Col{i+1}" for i, cell in enumerate(sheet[2])]

non_null_counts = {h: 0 for h in headers}

for r in range(3, sheet.max_row + 1):
    lead_id = sheet.cell(row=r, column=1).value
    if not lead_id or str(lead_id).strip() == "":
        continue
    for i, h in enumerate(headers):
        val = sheet.cell(row=r, column=i+1).value
        if val is not None and str(val).strip() not in ('', '-'):
            non_null_counts[h] += 1

print("Non-null counts in Excel (total active rows):")
for h, count in non_null_counts.items():
    print(f"  {h}: {count}")
