import openpyxl
from datetime import datetime

excel_path = '/Users/berkhan/Downloads/2026 - Mayıs Haziran Verileri.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Lead Takip']

records_by_month = {}
date_range_records = 0
min_date = None
max_date = None
valid_rows = 0

for r in range(3, sheet.max_row + 1):
    lead_id = sheet.cell(row=r, column=1).value
    if not lead_id: # Evaluates to empty or None
        continue
        
    valid_rows += 1
    date_val = sheet.cell(row=r, column=2).value # İlk Temas Tarihi
    
    # Try parsing date
    parsed_date = None
    if isinstance(date_val, datetime):
        parsed_date = date_val
    elif isinstance(date_val, str):
        # try standard parsing
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', '%d.%m.%Y', '%Y-%m-%d'):
            try:
                parsed_date = datetime.strptime(date_val.strip(), fmt)
                break
            except ValueError:
                pass
    
    if parsed_date:
        month_str = parsed_date.strftime('%Y-%m')
        records_by_month[month_str] = records_by_month.get(month_str, 0) + 1
        
        start_range = datetime(2026, 5, 8)
        end_range = datetime(2026, 6, 15, 23, 59, 59)
        if start_range <= parsed_date <= end_range:
            date_range_records += 1
            if min_date is None or parsed_date < min_date:
                min_date = parsed_date
            if max_date is None or parsed_date > max_date:
                max_date = parsed_date
    else:
        records_by_month['Unparsed/None'] = records_by_month.get('Unparsed/None', 0) + 1

print(f"Total valid rows with evaluated Lead ID: {valid_rows}")
print("Records by Month-Year:")
for k, v in sorted(records_by_month.items()):
    print(f"  {k}: {v}")
    
print(f"\nRecords between 08.05.2026 and 15.06.2026: {date_range_records}")
