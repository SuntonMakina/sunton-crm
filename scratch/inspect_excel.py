import openpyxl
from collections import Counter

excel_path = '/Users/berkhan/Downloads/2026 - Mayıs Haziran Verileri.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Lead Takip']

headers = [str(cell.value).strip() if cell.value is not None else f"Col{i+1}" for i, cell in enumerate(sheet[2])]

cols_to_inspect = [
    'Lead Durumu', 
    'Dönüş Olumlu mu?', 
    'Konuşma Yapıldı mı?', 
    'Satış Uzmanı', 
    'İletişim Kanalı',
    'Reklam Kaynağı'
]

print("Headers:", headers)
data = {col: [] for col in cols_to_inspect}

for r in range(3, sheet.max_row + 1):
    lead_id = sheet.cell(row=r, column=1).value
    if not lead_id or str(lead_id).strip() == "":
        continue
    for col in cols_to_inspect:
        if col in headers:
            col_idx = headers.index(col) + 1
            val = sheet.cell(row=r, column=col_idx).value
            data[col].append(val)

for col, vals in data.items():
    print(f"\n--- Unique values for '{col}' ---")
    c = Counter(vals)
    for val, count in c.most_common():
        print(f"  {val}: {count}")
