import openpyxl

excel_path = '/Users/berkhan/Downloads/2026 - Mayıs Haziran Verileri.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(row=r, column=c).value
            if val in (29, 28, 37, 9, 103, 71, 58, 116, 55):
                print(f"Sheet: {sheet_name} | Cell: {openpyxl.utils.get_column_letter(c)}{r} | Value: {val}")
