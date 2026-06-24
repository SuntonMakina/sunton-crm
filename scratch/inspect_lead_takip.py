import openpyxl

excel_path = '/Users/berkhan/Downloads/2026 - Mayıs Haziran Verileri.xlsx'
wb = openpyxl.load_workbook(excel_path)
sheet = wb['Lead Takip']

print("Sheet Max Row:", sheet.max_row)

# Row 2 contains headers
headers = [cell.value for cell in sheet[2]]
print("\nHeaders (Row 2):")
for i, h in enumerate(headers):
    print(f"Col {i+1}: {h}")

# Check data from row 3
data_rows = 0
valid_leads = 0
empty_rows = 0
for r in range(3, sheet.max_row + 1):
    lead_id = sheet.cell(row=r, column=1).value
    row_values = [sheet.cell(row=r, column=col).value for col in range(1, len(headers) + 1)]
    if any(val is not None for val in row_values):
        data_rows += 1
        if lead_id:
            valid_leads += 1
    else:
        empty_rows += 1

print(f"\nTotal data rows (at least one cell not empty): {data_rows}")
print(f"Rows with Lead ID: {valid_leads}")
print(f"Completely empty rows: {empty_rows}")
