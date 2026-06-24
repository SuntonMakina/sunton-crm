import openpyxl

excel_path = '/Users/berkhan/Downloads/2026 - Mayıs Haziran Verileri.xlsx'
wb = openpyxl.load_workbook(excel_path)
sheet = wb['Lead Takip']

count = 0
for r in range(3, sheet.max_row + 1):
    lead_id = sheet.cell(row=r, column=1).value
    if not lead_id:
        continue
    
    date_val = sheet.cell(row=r, column=2).value
    if date_val is None:
        count += 1
        row_values = [sheet.cell(row=r, column=col).value for col in range(1, 10)]
        print(f"Row {r}: {row_values}")
        if count >= 5:
            break
            
print(f"Total unparsed/missing date rows with Lead ID: {count}")
